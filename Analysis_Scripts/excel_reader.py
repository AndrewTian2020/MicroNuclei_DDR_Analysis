# ===================================================================================================
# Date: July 21, 2025 (Function originally written in July 11, 2025)
# Author: Andrew Tian
# This code was made to sort Kate's data and group it by well names and timepoints.
import re
import openpyxl
import string

# Defining a function to read the excel, extract wells, and extract mn ratios
def read_excel(path_to_excel):
    # Open the excel sheet
    workbook = openpyxl.load_workbook(path_to_excel)
    data_sheet = workbook.worksheets[0]

    # Extract all the column letters
    all_columns = list(string.ascii_uppercase)
    tempstring = ""

    for n in range(65, 91):
        tempstring = "A" + chr(n)
        all_columns.append(tempstring)

    for n in range(65, 83):
        tempstring = "B" + chr(n)
        all_columns.append(tempstring)

    # Initializing variables for overall well list and mn ratio lists
    well_list_plate_1 = []
    well_list_plate_2 = []

    mn_list_plate_1 = []
    mn_list_plate_2 = []

    # Initializing temp variables to store individual lists
    temp_names = []
    temp_names_plate_1 = []
    temp_names_plate_2 = []
    temp_mn = []
    temp_mn_plate_1 = []
    temp_mn_plate_2 = []
    well_mn = []
    match = ""
    i = 1
    n = 0
    a = 0

    # Loop through the excel and extract all important info
    for col in all_columns:

        # Every 5th column starting from the first contains image IDs
        if (i - 1) % 5 == 0:
            for val in data_sheet[col]:
                if val.value != "Image ID":
                    temp_names.append(val.value)

            while None in temp_names:
                temp_names.remove(None)

        if (i - 4) % 5 == 0:
            for num in data_sheet[col]:
                if num.value != "MN Ratio" and num.value != None:
                    temp_mn.append(float(num.value))
            
            while None in temp_mn:
                temp_mn.remove(None)
        i += 1

    # For loop to add names and mn ratios based on plates + wells
    for name in temp_names:
        if "Plate_1" in name:
            temp_names_plate_1.append(name)
            temp_mn_plate_1.append(temp_mn[a])
        else:
            temp_names_plate_2.append(name)
            temp_mn_plate_2.append(temp_mn[a])
        a += 1

    # For loop to add wells and mn ratios from plate 1
    for name in temp_names_plate_1:
        match = re.findall(r'GFP-H2B_([A-Z0-9]+)_', name)
        if match[0] not in well_list_plate_1 and len(well_mn) == 0:
            well_list_plate_1.append(match[0])
            well_mn.append(temp_mn_plate_1[n])
        # Second case where the well has already been added
        elif match[0] in well_list_plate_1:
            well_mn.append(temp_mn_plate_1[n])
        # Last case where the it starts on a new well after well 1
        else:
            mn_list_plate_1.append(well_mn)
            well_list_plate_1.append(match[0])
            well_mn = []
            well_mn.append(temp_mn_plate_1[n])
        n += 1
    mn_list_plate_1.append(well_mn)
    well_mn = []

    n = 0
    # For loop to add wells and mn ratios from plate 2
    for name in temp_names_plate_2:
        match = re.findall(r'GFP-H2B_([A-Z0-9]+)_', name)
        if match[0] not in well_list_plate_2 and len(well_mn) == 0:
            well_list_plate_2.append(match[0])
            well_mn.append(temp_mn_plate_2[n])
        # Second case where the well has already been added
        elif match[0] in well_list_plate_2:
            well_mn.append(temp_mn_plate_2[n])
        # Last case where the it starts on a new well after well 1
        else:
            mn_list_plate_2.append(well_mn)
            well_list_plate_2.append(match[0])
            well_mn = []
            well_mn.append(temp_mn_plate_2[n])
        n += 1
    mn_list_plate_2.append(well_mn)
    return well_list_plate_1, well_list_plate_2, mn_list_plate_1, mn_list_plate_2

# Read my outputted data sheet with outputted params/coefficients/data
def read_sheet(path_to_excel, n):
    # Take the workbook as input
    workbook = openpyxl.load_workbook(path_to_excel)
    worksheet = workbook.worksheets[n]

    # Initialize variables
    outputp1 = []
    outputp2 = []
    temp_list = []

    plate_1_col = ["A", "B", "C", "D", "E"]
    plate_2_col = ["G", "H", "I", "J", "K"]

    for col in plate_1_col:
        i = 0
        temp_list = []
        for val in worksheet[col]:
            if i != 0 and val.value != None:
                try:
                    temp_list.append(float(val.value))
                except:
                    temp_list.append(val.value)
            i += 1
        outputp1.append(temp_list)
    
    for col in plate_2_col:
        i = 0
        temp_list = []
        for val in worksheet[col]:
            if i != 0 and val.value != None:
                try:
                    temp_list.append(float(val.value))
                except:
                    temp_list.append(val.value)
            i += 1
        outputp2.append(temp_list)

    return outputp1, outputp2

# Read my outputted data sheet with outputted xmaxes + mn maxes + correlation
def read_sheet2(path_to_excel, n):
    # Take the workbook as input
    workbook = openpyxl.load_workbook(path_to_excel)
    worksheet = workbook.worksheets[n]

    # Initialize variables
    outputp1 = []
    outputp2 = []
    temp_list = []

    plate_1_col = ["A", "B", "C", "D"]
    plate_2_col = ["F", "G", "H", "I"]

    for col in plate_1_col:
        i = 0
        temp_list = []
        for val in worksheet[col]:
            if i != 0 and val.value != None:
                try:
                    temp_list.append(float(val.value))
                except:
                    temp_list.append(val.value)
            i += 1
        outputp1.append(temp_list)
    
    for col in plate_2_col:
        i = 0
        temp_list = []
        for val in worksheet[col]:
            if i != 0 and val.value != None:
                try:
                    temp_list.append(float(val.value))
                except:
                    temp_list.append(val.value)
            i += 1
        outputp2.append(temp_list)

    return outputp1, outputp2

def extract_well_map(path_to_excel, n):
    dict_map = {}
    well_list = []
    gene_list = []
    workbook = openpyxl.load_workbook(path_to_excel)
    worksheet = workbook.worksheets[n]

    for val in worksheet["A"]:
        well_list.append(val.value)
    for val in worksheet["B"]:
        gene_list.append(val.value)
    
    del well_list[:2]
    del gene_list[:2]

    for well in well_list:
        i = well_list.index(well)
        if well not in dict_map:
            dict_map[well] = gene_list[i]
    return dict_map

def extract_DDR_map(path_to_excel, n):
    DDR_map = {}
    workbook = openpyxl.load_workbook(path_to_excel)
    worksheet = workbook.worksheets[n]

    cols = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]

    for col in cols:
        temp_list = []
        temp_path = ""
        path_list = []
        for val in worksheet[col]:
            if val.value != None:
                temp_list.append(val.value)
        i = 0
        for item in temp_list:
            if i == 0:
                temp_path = item
            else:
                path_list.append(item)
            i += 1
        DDR_map[temp_path] = path_list
    return DDR_map